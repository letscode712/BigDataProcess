#!/usr/bin/python3

import openpyxl

fpath = "C:\Users\kimyo\OneDrive\문서\3학년, 2023-2학기\04_빅데이터처리\HW1\student.xlsx"
wb = openpyxl.load_workbook(fpath)
st = wb.active ##sheet1을 st로 불러옴

for row in range(2, 76):
    st.cell(row = row, column='total', value= st['midterm']*0.30+st['final']**0.35+st['hw']*0.34+st['attendance']*0.01)
    
st_list = [st['total'].sort()]
Amax = int(74*0.3)
Bmax = int(74*0.7)

for row in range(2,76):
    total = st.cell(row=row, column='total').value
    if total < 40.00:
        grade = 'F'
    elif total >= st_list[Amax-1]:
        if total >= st_list[Amax + int(Amax*0.5) -1]:
            grade = 'A+'
        grade = 'A0'
    elif total >= st_list[Bmax-1]:
        if total >= st_list[Bmax + int((74-Bmax)*0.5)-1]:
            grade = 'B+'
        grade = 'B0'
    elif total >= st_list[73]:
        if total >= st_list[74-int((74-Bmax)*0.5)-1]:
            grade = 'C+'
        grade = 'C0'
    grade = 'F'
    st.cell(row=row, column='grade', value = grade)
    
wb.save(fpath)
